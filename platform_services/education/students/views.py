from platform_services.identity.mixins import TenantQuerySetMixin
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.core.files.storage import default_storage
import openpyxl
from datetime import datetime
import unicodedata


from .models import Student, Attendance, Enrollment
from .serializers import StudentSerializer, AttendanceSerializer, EnrollmentSerializer

class EnrollmentViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    filterset_fields = ['student', 'academic_year', 'school_class', 'decision']


class StudentViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filterset_fields = ['lifecycle_status', 'is_archived']

    def get_queryset(self):
        queryset = Student.objects.prefetch_related('enrollments', 'enrollments__school_class').all()
        is_archived = self.request.query_params.get('is_archived', 'false')
        if is_archived.lower() == 'true':
            queryset = queryset.filter(is_archived=True)
        else:
            queryset = queryset.filter(is_archived=False)
            
        school_class = self.request.query_params.get('school_class')
        if school_class:
            queryset = queryset.filter(enrollments__school_class_id=school_class).distinct()
            
        return queryset

    def perform_create(self, serializer):
        student = serializer.save()
        school_class_id = self.request.data.get('school_class')
        if school_class_id:
            from platform_services.education.classes.models import AcademicYear
            from .models import Enrollment
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                Enrollment.objects.create(
                    student=student,
                    academic_year=active_year,
                    school_class_id=school_class_id,
                    decision='EN_COURS'
                )

    def perform_update(self, serializer):
        student = serializer.save()
        school_class_id = self.request.data.get('school_class')
        if school_class_id:
            from platform_services.education.classes.models import AcademicYear
            from .models import Enrollment
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                enrollment, _ = Enrollment.objects.get_or_create(
                    student=student,
                    academic_year=active_year,
                    defaults={'school_class_id': school_class_id, 'decision': 'EN_COURS'}
                )
                if enrollment.school_class_id != int(school_class_id):
                    enrollment.school_class_id = school_class_id
                    enrollment.save()

    def perform_destroy(self, instance):
        instance.is_archived = True
        instance.save()

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        student = self.get_object()
        student.is_archived = False
        student.save()
        return Response({'status': 'student restored'})
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Le fichier doit être au format .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            created = 0
            errors = []
            
            # Read headers
            headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
            
            # Map headers to indices
            header_map = {
                'last_name': -1, 'first_name': -1, 'sex': -1, 'date_of_birth': -1,
                'place_of_birth': -1, 'school_class': -1, 'parent_name': -1, 'parent_phone': -1
            }
            for idx, header in enumerate(headers):
                header_norm = unicodedata.normalize('NFKD', header).encode('ASCII', 'ignore').decode('utf-8')
                if 'nom' in header_norm and 'parent' not in header_norm: header_map['last_name'] = idx
                elif 'prenom' in header_norm: header_map['first_name'] = idx
                elif 'sexe' in header_norm or 'genre' in header_norm: header_map['sex'] = idx
                elif 'date' in header_norm and 'naiss' in header_norm: header_map['date_of_birth'] = idx
                elif 'lieu' in header_norm and 'naiss' in header_norm: header_map['place_of_birth'] = idx
                elif 'classe' in header_norm: header_map['school_class'] = idx
                elif 'parent' in header_norm and ('nom' in header_norm or 'tuteur' in header_norm): header_map['parent_name'] = idx
                elif 'tel' in header_norm or 'phone' in header_norm: header_map['parent_phone'] = idx
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or (header_map['last_name'] != -1 and not row[header_map['last_name']]): 
                    if row and row[0]: # Try fallback
                        pass
                    else:
                        continue
                
                def get_val(key, default=''):
                    idx = header_map[key]
                    if idx != -1 and idx < len(row) and row[idx] is not None:
                        return row[idx]
                    return default
                
                last_name = str(get_val('last_name'))
                if not last_name:
                    # If we couldn't map last_name, fallback to first column
                    if row[0]:
                        last_name = str(row[0])
                    else:
                        errors.append(f"Ligne {row_idx}: Nom manquant.")
                        continue
                        
                first_name = str(get_val('first_name'))
                sex = str(get_val('sex', 'M'))[:1].upper()
                if sex not in ['M', 'F']: sex = 'M'
                
                dob_raw = get_val('date_of_birth')
                if isinstance(dob_raw, datetime):
                    dob = dob_raw.date()
                elif dob_raw:
                    try:
                        dob = datetime.strptime(str(dob_raw).strip(), "%Y-%m-%d").date()
                    except:
                        dob = datetime.strptime("2000-01-01", "%Y-%m-%d").date()
                else:
                    dob = datetime.strptime("2000-01-01", "%Y-%m-%d").date()
                        
                place_of_birth = str(get_val('place_of_birth', 'Inconnu'))
                school_class_raw = get_val('school_class')
                parent_name = str(get_val('parent_name', 'Inconnu'))
                parent_phone = str(get_val('parent_phone'))
                
                school_class_id = None
                if school_class_raw:
                    from platform_services.education.classes.models import SchoolClass
                    if str(school_class_raw).strip().isdigit():
                        school_class = SchoolClass.objects.filter(id=int(school_class_raw)).first()
                    else:
                        school_class = SchoolClass.objects.filter(name__iexact=str(school_class_raw).strip()).first()
                        
                    if school_class:
                        school_class_id = school_class.id
                    else:
                        errors.append(f"Ligne {row_idx}: Classe '{school_class_raw}' introuvable.")
                
                try:
                    student = Student(
                        last_name=last_name,
                        first_name=first_name,
                        sex=sex,
                        date_of_birth=dob,
                        place_of_birth=place_of_birth,
                        parent_name=parent_name,
                        parent_phone=parent_phone
                    )
                    student.save()
                    
                    if school_class_id:
                        from platform_services.education.classes.models import AcademicYear
                        from platform_services.education.students.models import Enrollment
                        active_year = AcademicYear.objects.filter(is_active=True).first()
                        if active_year:
                            Enrollment.objects.get_or_create(
                                student=student,
                                academic_year=active_year,
                                school_class_id=school_class_id,
                                defaults={'decision': 'EN_COURS'}
                            )
                            
                    created += 1
                except Exception as e:
                    errors.append(f"Ligne {row_idx}: Erreur lors de l'intégration ({str(e)})")
                    
            return Response({'message': f'{created} élèves importés avec succès.', 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class AttendanceViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    filterset_fields = ['student', 'date', 'academic_year', 'sequence']
    
    @action(detail=False, methods=['post'])
    def batch(self, request):
        attendances_data = request.data.get('attendances', [])
        # expects: [{'student': ID, 'date': 'YYYY-MM-DD', 'academic_year': ID, 'sequence': str, 'is_absent': bool}]
        
        if not attendances_data:
            return Response({'error': 'No data provided.'}, status=status.HTTP_400_BAD_REQUEST)
            
        # Check CalendarEngine for the first date in batch (assuming batch is usually for same date)
        sample_date = attendances_data[0].get('date')
        if sample_date:
            from platform_services.education.classes.services import CalendarEngine
            if CalendarEngine.is_attendance_suspended(sample_date, organization_id=request.user.organization_id if hasattr(request.user, 'organization_id') else None):
                return Response({'error': 'L\'appel est suspendu pour cette date (jour férié ou événement spécial).'}, status=status.HTTP_400_BAD_REQUEST)
        
        created = 0
        updated = 0
        
        for a_data in attendances_data:
            student_id = a_data.get('student')
            date = a_data.get('date')
            sequence = a_data.get('sequence')
            academic_year_id = a_data.get('academic_year')
            is_absent = a_data.get('is_absent', False)
            
            if not all([student_id, date, sequence, academic_year_id]):
                continue
                
            att_obj, is_new = Attendance.objects.get_or_create(
                student_id=student_id,
                date=date,
                sequence=sequence,
                defaults={
                    'academic_year_id': academic_year_id,
                    'is_absent': is_absent,
                    'organization_id': request.user.organization_id if hasattr(request.user, 'organization_id') else None
                }
            )
            
            if not is_new:
                att_obj.is_absent = is_absent
                att_obj.save()
                updated += 1
            else:
                created += 1
                
        return Response({'message': 'Appel enregistré.', 'created': created, 'updated': updated})

